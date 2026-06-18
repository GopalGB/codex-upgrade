#!/usr/bin/env python3
"""xlsx.py - memory-safe Excel toolkit for BIG .xlsx files (no MCP, no pandas needed).

Reads use openpyxl's read_only mode (constant memory, streams row-by-row), so a
500 MB workbook does not blow up RAM. Writes use xlsxwriter constant_memory mode.

Subcommands:
  info   FILE                         sheets, dimensions, on-disk size
  head   FILE [--sheet S] [--n 20]    first N rows of a sheet (TSV)
  tail   FILE [--sheet S] [--n 20]    last N rows (single streamed pass)
  to-csv FILE [--sheet S] [-o OUT]    stream a sheet to CSV (memory-safe)
  stats  FILE [--sheet S]             per-column type/null/min/max/sum (streamed)
  grep   FILE PATTERN [--sheet S]     rows containing a regex match (streamed)
  from-csv OUT.xlsx --csv IN.csv [--sheet name]   write CSV -> xlsx (constant memory)

Examples:
  python xlsx.py info big.xlsx
  python xlsx.py to-csv big.xlsx --sheet Data -o data.csv
  python xlsx.py stats big.xlsx --sheet Data
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from pathlib import Path

_HERE = os.path.dirname(os.path.abspath(__file__))
for _c in (
    os.path.join(_HERE, "..", "..", "..", "lib"),
    os.path.expanduser("~/.codex/lib"),
    os.path.expanduser("~/.agents/lib"),
):
    if os.path.isfile(os.path.join(_c, "codex_env.py")):
        sys.path.insert(0, os.path.abspath(_c))
        break
import codex_env  # noqa: E402


def _human(n: int) -> str:
    for unit in ("B", "KB", "MB", "GB", "TB"):
        if n < 1024 or unit == "TB":
            return (
                f"{n:.0f}{unit}"
                if unit == "B"
                else f"{n / 1:.0f}{unit}"
                if False
                else f"{n:.1f}{unit}"
            )
        n /= 1024
    return f"{n:.1f}TB"


def _open_ro(path: str):
    import openpyxl

    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _pick_sheet(wb, name: str | None):
    if name is None:
        return wb[wb.sheetnames[0]]
    if name not in wb.sheetnames:
        sys.exit(f"sheet '{name}' not found. Available: {', '.join(wb.sheetnames)}")
    return wb[name]


def cmd_info(a):
    size = Path(a.file).stat().st_size
    wb = _open_ro(a.file)
    print(f"file:  {a.file}  ({_human(size)})")
    print(f"sheets: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"  - {name}: {ws.max_row or '?'} rows x {ws.max_column or '?'} cols")
    wb.close()


def cmd_head(a):
    wb = _open_ro(a.file)
    ws = _pick_sheet(wb, a.sheet)
    out = csv.writer(sys.stdout, delimiter="\t")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= a.n:
            break
        out.writerow(["" if c is None else c for c in row])
    wb.close()


def cmd_tail(a):
    from collections import deque

    wb = _open_ro(a.file)
    ws = _pick_sheet(wb, a.sheet)
    buf = deque(maxlen=a.n)
    for row in ws.iter_rows(values_only=True):
        buf.append(row)
    out = csv.writer(sys.stdout, delimiter="\t")
    for row in buf:
        out.writerow(["" if c is None else c for c in row])
    wb.close()


def cmd_to_csv(a):
    wb = _open_ro(a.file)
    ws = _pick_sheet(wb, a.sheet)
    fh = open(a.output, "w", newline="", encoding="utf-8") if a.output else sys.stdout
    try:
        out = csv.writer(fh)
        n = 0
        for row in ws.iter_rows(values_only=True):
            out.writerow(["" if c is None else c for c in row])
            n += 1
        if a.output:
            print(f"wrote {n} rows -> {a.output}", file=sys.stderr)
    finally:
        if a.output:
            fh.close()
        wb.close()


def cmd_stats(a):
    wb = _open_ro(a.file)
    ws = _pick_sheet(wb, a.sheet)
    headers, cols = None, []
    rows = 0
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [
                str(c) if c is not None else f"col{i}" for i, c in enumerate(row)
            ]
            cols = [
                {"n": 0, "nulls": 0, "nums": 0, "min": None, "max": None, "sum": 0.0}
                for _ in headers
            ]
            continue
        rows += 1
        for i, val in enumerate(row):
            if i >= len(cols):
                break
            c = cols[i]
            c["n"] += 1
            if val is None or val == "":
                c["nulls"] += 1
                continue
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                c["nums"] += 1
                c["sum"] += float(val)
                c["min"] = val if c["min"] is None else min(c["min"], val)
                c["max"] = val if c["max"] is None else max(c["max"], val)
    wb.close()
    print(f"rows: {rows}  cols: {len(headers or [])}")
    print(f"{'column':<28}{'type':<10}{'nulls':>8}{'min':>14}{'max':>14}{'sum':>16}")
    for h, c in zip(headers or [], cols):
        kind = (
            "numeric"
            if c["nums"] >= max(1, c["n"] - c["nulls"]) and c["nums"]
            else "text"
        )
        mn = "" if c["min"] is None else f"{c['min']}"
        mx = "" if c["max"] is None else f"{c['max']}"
        sm = f"{c['sum']:.2f}" if c["nums"] else ""
        print(f"{h[:27]:<28}{kind:<10}{c['nulls']:>8}{mn:>14}{mx:>14}{sm:>16}")


def cmd_grep(a):
    rx = re.compile(a.pattern)
    wb = _open_ro(a.file)
    ws = _pick_sheet(wb, a.sheet)
    out = csv.writer(sys.stdout, delimiter="\t")
    hits = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(c is not None and rx.search(str(c)) for c in row):
            out.writerow([idx] + ["" if c is None else c for c in row])
            hits += 1
            if a.max and hits >= a.max:
                break
    wb.close()
    print(f"{hits} matching row(s)", file=sys.stderr)


def cmd_from_csv(a):
    import xlsxwriter

    wbk = xlsxwriter.Workbook(a.output, {"constant_memory": True})
    sh = wbk.add_worksheet(a.sheet[:31])
    with open(a.csv, newline="", encoding="utf-8") as fh:
        for r, row in enumerate(csv.reader(fh)):
            for c, val in enumerate(row):
                # try to coerce numbers so Excel treats them as numbers
                try:
                    sh.write_number(
                        r,
                        c,
                        float(val) if "." in val or "e" in val.lower() else int(val),
                    )
                except (ValueError, OverflowError):
                    sh.write_string(r, c, val)
    wbk.close()
    print(f"wrote {a.output}")


def main():
    p = argparse.ArgumentParser(description="memory-safe big-xlsx toolkit")
    sub = p.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("info")
    s.add_argument("file")
    s.set_defaults(fn=cmd_info, deps=["openpyxl"])
    for name, fn in (("head", cmd_head), ("tail", cmd_tail)):
        s = sub.add_parser(name)
        s.add_argument("file")
        s.add_argument("--sheet")
        s.add_argument("--n", type=int, default=20)
        s.set_defaults(fn=fn, deps=["openpyxl"])
    s = sub.add_parser("to-csv")
    s.add_argument("file")
    s.add_argument("--sheet")
    s.add_argument("-o", "--output")
    s.set_defaults(fn=cmd_to_csv, deps=["openpyxl"])
    s = sub.add_parser("stats")
    s.add_argument("file")
    s.add_argument("--sheet")
    s.set_defaults(fn=cmd_stats, deps=["openpyxl"])
    s = sub.add_parser("grep")
    s.add_argument("file")
    s.add_argument("pattern")
    s.add_argument("--sheet")
    s.add_argument("--max", type=int, default=0)
    s.set_defaults(fn=cmd_grep, deps=["openpyxl"])
    s = sub.add_parser("from-csv")
    s.add_argument("output")
    s.add_argument("--csv", required=True)
    s.add_argument("--sheet", default="Sheet1")
    s.set_defaults(fn=cmd_from_csv, deps=["xlsxwriter"])

    a = p.parse_args()
    codex_env.bootstrap(*a.deps)
    a.fn(a)


if __name__ == "__main__":
    main()
